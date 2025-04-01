import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")

        # Create a menu bar
        self.menu_bar = tk.Menu(self.root)
        self.root.config(menu=self.menu_bar)

        # Create the File menu
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)
        self.file_menu.add_command(label="New File", command=self.create_new_file)
        self.file_menu.add_command(label="Open File", command=self.load_file)
        self.file_menu.add_command(label="Save File", command=self.save_file)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Exit", command=self.root.quit)

        # Load last selected month
        self.config_file = "config.txt"
        self.last_month = self.load_last_month()

        # Create sections
        self.date_section = tk.Frame(self.root)
        self.date_section.pack(fill="x")

        self.category_section = tk.Frame(self.root)
        self.category_section.pack(fill="x")

        self.amount_section = tk.Frame(self.root)
        self.amount_section.pack(fill="x")

        self.custom_category_section = tk.Frame(self.root)
        self.custom_category_section.pack(fill="x")

        # Create entry fields
        self.date_label = tk.Label(self.date_section, text="Date (DD):")
        self.date_label.pack(side="left")
        self.date_entry = tk.Entry(self.date_section, width=5)
        self.date_entry.pack(side="left")
        self.date_entry.bind("<Return>", lambda event: self.category_menu.focus())

        self.month_label = tk.Label(self.date_section, text="Month:")
        self.month_label.pack(side="left")
        self.month_var = tk.StringVar()
        self.month_var.set(self.last_month)  # Set to last changed month
        self.month_menu = ttk.Combobox(self.date_section, textvariable=self.month_var, values=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
        self.month_menu.pack(side="left")
        self.month_menu.bind("<<ComboboxSelected>>", self.save_last_month)

        self.year_label = tk.Label(self.date_section, text="Year:")
        self.year_label.pack(side="left")
        self.year_var = tk.StringVar()
        self.year_var.set("2024")  # default year
        self.year_menu = ttk.Combobox(self.date_section, textvariable=self.year_var, values=["2023", "2024", "2025"])
        self.year_menu.pack(side="left")

        self.category_label = tk.Label(self.category_section, text="Category:")
        self.category_label.pack(side="left")
        self.category_var = tk.StringVar()
        self.category_var.set("Select Category")  # default category
        self.category_menu = ttk.Combobox(self.category_section, textvariable=self.category_var, values=["Select Category"])
        self.category_menu.pack(side="left")
        self.category_menu.bind("<Return>", lambda event: self.amount_entry.focus())
        self.category_menu.bind("<<ComboboxSelected>>", self.focus_on_amount)

        self.add_category_button = tk.Button(self.category_section, text="Add Category", command=self.show_custom_category_entry)
        self.add_category_button.pack(side="left")

        self.custom_category_label = tk.Label(self.custom_category_section, text="New Category:")
        self.custom_category_label.pack(side="left")
        self.custom_category_entry = tk.Entry(self.custom_category_section)
        self.custom_category_entry.pack(side="left")
        self.custom_category_entry.bind("<Return>", self.add_category)
        self.custom_category_section.pack_forget()

        self.amount_label = tk.Label(self.amount_section, text="Amount:")
        self.amount_label.pack(side="left")
        self.amount_entry = tk.Entry(self.amount_section)
        self.amount_entry.pack(side="left")
        self.amount_entry.bind("<Return>", self.submit_form)

        # Create a button to submit the form
        self.submit_button = tk.Button(self.root, text="Submit", command=self.submit_form)
        self.submit_button.pack(fill="x")

        # Initialize workbook and filename
        self.wb = None
        self.file_name = "expenses.xlsx"
        self.load_or_create_file(self.file_name)

        # Create a dictionary to store the data month-wise and category-wise
        self.monthly_expenses = {}
        self.category_expenses = {}

    def load_last_month(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as file:
                return file.read().strip()
        return "January"

    def save_last_month(self, event=None):
        with open(self.config_file, "w") as file:
            file.write(self.month_var.get())

    def get_month_number(self, month_name):
        month_numbers = {
            "January": 1,
            "February": 2,
            "March": 3,
            "April": 4,
            "May": 5,
            "June": 6,
            "July": 7,
            "August": 8,
            "September": 9,
            "October": 10,
            "November": 11,
            "December": 12
        }
        return month_numbers[month_name]

    def show_custom_category_entry(self):
        self.custom_category_section.pack(fill="x")
        self.custom_category_entry.focus()

    def add_category(self, event=None):
        category = self.custom_category_entry.get()
        if category and category not in self.category_menu['values']:
            self.category_menu['values'] = (*self.category_menu['values'], category)
            self.category_var.set(category)
            self.custom_category_entry.delete(0, tk.END)
            self.custom_category_section.pack_forget()
            self.focus_on_amount()

    def focus_on_amount(self, event=None):
        self.amount_entry.focus()

    def load_or_create_file(self, file_name):
        if os.path.exists(file_name):
            self.wb = load_workbook(file_name)
        else:
            self.wb = Workbook()
            self.wb.save(file_name)
        self.file_name = file_name

    def create_new_file(self):
        new_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if new_file:
            self.wb = Workbook()
            self.wb.save(new_file)
            self.file_name = new_file

    def load_file(self):
        file_to_open = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_to_open:
            self.load_or_create_file(file_to_open)

    def save_file(self):
        if self.file_name:
            self.wb.save(self.file_name)

    def submit_form(self, event=None):
        # Get the values from the entry fields
        day = self.date_entry.get()
        month = self.month_var.get()
        year = self.year_var.get()
        date_str = f"{year}-{self.get_month_number(month)}-{day}"
        category = self.category_var.get()
        amount = float(self.amount_entry.get())

        # Parse the date string to get the month
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_name = date_obj.strftime("%B")

        # Add the data to the dictionary
        if month_name not in self.monthly_expenses:
            self.monthly_expenses[month_name] = []
        self.monthly_expenses[month_name].append((category, amount))

        if category not in self.category_expenses:
            self.category_expenses[category] = []
        self.category_expenses[category].append((month_name, amount))

        # Add the data to the corresponding category sheet
        if category not in self.wb.sheetnames:
            ws = self.wb.create_sheet(title=category)
            ws.append(["Date", "Amount"])
        else:
            ws = self.wb[category]

        ws.append([date_str, amount])

        # Update the monthly total for the category
        monthly_total = sum(amount for m, amount in self.category_expenses[category] if m == month_name)
        if "Totals" not in self.wb.sheetnames:
            total_ws = self.wb.create_sheet(title="Totals")
            total_ws.append(["Category", "Month", "Total"])
        else:
            total_ws = self.wb["Totals"]

        # Update or add the monthly total
        updated = False
        for row in range(2, total_ws.max_row + 1):
            if (total_ws.cell(row=row, column=1).value == category and 
                total_ws.cell(row=row, column=2).value == month_name):
                total_ws.cell(row=row, column=3, value=monthly_total)
                updated = True
                break
        if not updated:
            total_ws.append([category, month_name, monthly_total])

        # Clear the entry fields and focus back to date entry
        self.date_entry.delete(0, tk.END)
        self.category_var.set("Select Category")
        self.amount_entry.delete(0, tk.END)
        self.date_entry.focus()

        # Save the workbook
        self.save_file()

        print(f"Entry added: {date_str}, {category}, {amount}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
