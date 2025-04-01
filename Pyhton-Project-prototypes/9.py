import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
import os

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")
        self.root.geometry("500x400")
        self.root.configure(bg="black")
        
        self.file_name = "expenses.xlsx"
        self.load_or_create_file(self.file_name)
        
        # Apply a theme
        style = ttk.Style()
        style.theme_use('clam')
        
        # Custom Fonts and Colors
        label_font = ("Helvetica", 12, "bold")
        button_font = ("Helvetica", 14, "bold")
        entry_bg = "white"
        entry_fg = "black"
        entry_width = 20

        # Create a menu bar
        self.menu_bar = tk.Menu(self.root, bg="#264653", fg="white")
        self.root.config(menu=self.menu_bar)
        
        # Create the File menu
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0, bg="#e9c46a", fg="black")
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)
        self.file_menu.add_command(label="New File", command=self.create_new_file)
        self.file_menu.add_command(label="Exit", command=self.root.quit)
        
        # Create Status Bar
        self.status_bar = tk.Label(self.root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W, bg="#264653", fg="white")
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Form layout
        form_frame = tk.Frame(self.root, bg="black")
        form_frame.pack(padx=10, pady=10, anchor="w")
        
        self.date_label = tk.Label(form_frame, text="Date (DD-MM-YYYY):", font=label_font, fg="white", bg="black")
        self.date_label.grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.date_entry = ttk.Entry(form_frame, width=entry_width)
        self.date_entry.grid(row=0, column=1, padx=5, pady=5)
        
        self.category_label = tk.Label(form_frame, text="Category:", font=label_font, fg="white", bg="black")
        self.category_label.grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.category_var = tk.StringVar()
        self.category_menu = ttk.Combobox(form_frame, textvariable=self.category_var, values=["Food", "Transport", "Rent", "Entertainment", "Utilities", "Miscellaneous"], width=entry_width-1)
        self.category_menu.grid(row=1, column=1, padx=5, pady=5)
        
        self.amount_label = tk.Label(form_frame, text="Amount:", font=label_font, fg="white", bg="black")
        self.amount_label.grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.amount_entry = ttk.Entry(form_frame, width=entry_width)
        self.amount_entry.grid(row=2, column=1, padx=5, pady=5)
        
        self.submit_button = tk.Button(self.root, text="Submit", font=button_font, bg="blue", fg="white", command=self.add_expense)
        self.submit_button.pack(pady=5)
        
        self.view_button = tk.Button(self.root, text="View Expenses", font=button_font, bg="green", fg="white", command=self.view_expenses)
        self.view_button.pack(pady=5)
        
        self.expenses_list = ttk.Treeview(self.root, columns=("date", "category", "amount"), show="headings")
        self.expenses_list.heading("date", text="Date")
        self.expenses_list.heading("category", text="Category")
        self.expenses_list.heading("amount", text="Amount")
        self.expenses_list.pack()
        
    def load_or_create_file(self, file_name):
        if os.path.exists(file_name):
            self.wb = load_workbook(file_name)
            self.update_status_bar(f"File loaded: {os.path.basename(file_name)}")
        else:
            self.wb = Workbook()
            self.wb.save(file_name)
            self.update_status_bar(f"New file created: {os.path.basename(file_name)}")
    
    def create_new_file(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if file_path:
            self.wb = Workbook()
            self.wb.save(file_path)
            self.file_name = file_path
            self.update_status_bar(f"New file created: {os.path.basename(file_path)}")
    
    def add_expense(self):
        date = self.date_entry.get()
        category = self.category_var.get()
        amount = self.amount_entry.get()
        if date and category and amount:
            if category not in self.wb.sheetnames:
                ws = self.wb.create_sheet(title=category)
                ws.append(["Date", "Month", "Amount", "Remark"])
            else:
                ws = self.wb[category]
            
            ws.append([date, date.split('-')[1], int(amount), ""])
            self.wb.save(self.file_name)
            
            self.view_expenses()
            messagebox.showinfo("Success", "Expense added successfully!")
            self.update_status_bar("Expense added!")
            self.date_entry.delete(0, tk.END)
            self.category_menu.set("")
            self.amount_entry.delete(0, tk.END)
            self.date_entry.focus()
        else:
            messagebox.showerror("Error", "Please enter all fields")
    
    def view_expenses(self):
        for row in self.expenses_list.get_children():
            self.expenses_list.delete(row)
        
        for sheet in self.wb.sheetnames:
            ws = self.wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                self.expenses_list.insert("", "end", values=row[:3])
    
    def update_status_bar(self, message):
        self.status_bar.config(text=message)
    
if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
