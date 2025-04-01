import tkinter as tk
import sys
from PyQt5.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QLineEdit, QTableWidget, QTableWidgetItem, QAbstractItemView
from PyQt5.QtCore import Qt
import openpyxl
from openpyxl.utils import get_column_letter

class ExpenseTracker(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.setGeometry(300, 300, 800, 600)
        self.setWindowTitle('Monthly Expense Tracker')

        self.layout = QVBoxLayout()
        self.setLayout(self.layout)

        # Input fields
        self.dateEdit = QLineEdit()
        self.dateEdit.setPlaceholderText('Date (DD/MM/YYYY)')
        self.layout.addWidget(self.dateEdit)

        self.categoryEdit = QLineEdit()
        self.categoryEdit.setPlaceholderText('Category (e.g. Food, Transport, etc.)')
        self.layout.addWidget(self.categoryEdit)

        self.amountEdit = QLineEdit()
        self.amountEdit.setPlaceholderText('Amount')
        self.layout.addWidget(self.amountEdit)

        # Add button
        addButton = QPushButton('Add Expense')
        addButton.clicked.connect(self.addExpense)
        self.layout.addWidget(addButton)

        # Table to display expenses
        self.expenseTable = QTableWidget()
        self.expenseTable.setRowCount(0)
        self.expenseTable.setColumnCount(3)
        self.expenseTable.setHorizontalHeaderLabels(['Date', 'Category', 'Amount'])
        self.expenseTable.horizontalHeader().setStretchLastSection(True)
        self.expenseTable.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.layout.addWidget(self.expenseTable)

        # Load existing expenses from Excel sheet
        self.loadExpenses()

    def addExpense(self):
        date = self.dateEdit.text()
        category = self.categoryEdit.text()
        amount = self.amountEdit.text()

        if date and category and amount:
            # Add expense to Excel sheet
            self.addExpenseToExcel(date, category, amount)

            # Clear input fields
            self.dateEdit.clear()
            self.categoryEdit.clear()
            self.amountEdit.clear()

            # Update table
            self.updateTable()

    def loadExpenses(self):
        # Load existing expenses from Excel sheet
        wb = openpyxl.load_workbook('expenses.xlsx')
        sheet = wb.active

        for row in sheet.iter_rows(values_only=True):
            self.expenseTable.insertRow(0)
            for col, value in enumerate(row):
                item = QTableWidgetItem(str(value))
                self.expenseTable.setItem(0, col, item)

    def addExpenseToExcel(self, date, category, amount):
        # Add expense to Excel sheet
        wb = openpyxl.load_workbook('expenses.xlsx')
        sheet = wb.active

        row = [date, category, amount]
        sheet.append(row)

        wb.save('expenses.xlsx')

    def updateTable(self):
        # Update table with new expenses
        self.expenseTable.setRowCount(0)
        self.loadExpenses()

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ExpenseTracker()
    window.show()
    sys.exit(app.exec_())


class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")

        # Create sections
        self.date_section = tk.Frame(self.root)
        self.date_section.pack(fill="x")

        self.category_section = tk.Frame(self.root)
        self.category_section.pack(fill="x")

        self.amount_section = tk.Frame(self.root)
        self.amount_section.pack(fill="x")

        # Create entry fields
        self.date_label = tk.Label(self.date_section, text="Date:")
        self.date_label.pack(side="left")
        self.date_entry = tk.Entry(self.date_section)
        self.date_entry.pack(side="left")
        self.date_entry.bind("<Return>", lambda event: self.category_entry.focus())

        self.category_label = tk.Label(self.category_section, text="Category:")
        self.category_label.pack(side="left")
        self.category_entry = tk.Entry(self.category_section)
        self.category_entry.pack(side="left")
        self.category_entry.bind("<Return>", lambda event: self.amount_entry.focus())

        self.amount_label = tk.Label(self.amount_section, text="Amount:")
        self.amount_label.pack(side="left")
        self.amount_entry = tk.Entry(self.amount_section)
        self.amount_entry.pack(side="left")
        self.amount_entry.bind("<Return>", lambda event: self.root.focus())  # Move focus to root window

        # Add a button to submit the form
        self.submit_button = tk.Button(self.root, text="Submit", command=self.submit_form)
        self.submit_button.pack(fill="x")

    def submit_form(self):
        # Get the values from the entry fields
        date = self.date_entry.get()
        category = self.category_entry.get()
        amount = self.amount_entry.get()

        # Do something with the values (e.g., save to a database or file)
        print(f"Date: {date}, Category: {category}, Amount: {amount}")

root = tk.Tk()
app = ExpenseTracker(root)
root.mainloop()
