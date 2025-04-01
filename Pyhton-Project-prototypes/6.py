import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")
        self.root.geometry("500x400")
        self.root.configure(bg="#f4a261")

        # Apply a theme
        style = ttk.Style()
        style.theme_use('clam')

        # Custom Fonts and Colors
        label_font = ("Helvetica", 12)
        button_font = ("Helvetica", 14, "bold")
        entry_bg = "#fff8e1"

        # Create a menu bar
        self.menu_bar = tk.Menu(self.root)
        self.root.config(menu=self.menu_bar)

        # Create the File menu
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0, bg="#e9c46a")
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)
        self.file_menu.add_command(label="New File", command=self.create_new_file)
        self.file_menu.add_command(label="Open File", command=self.load_file)
        self.file_menu.add_command(label="Save File", command=self.save_file)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Exit", command=self.root.quit)

        # Create Status Bar (early initialization)
        self.status_bar = tk.Label(self.root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W, bg="#264653", fg="white")
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Load last selected month
        self.config_file = "config.txt"
        self.last_month = self.load_last_month()

        # Create sections
        self.date_section = tk.Frame(self.root, bg="#f4a261")
        self.date_section.pack(fill="x", padx=10, pady=5)

        self.category_section = tk.Frame(self.root, bg="#f4a261")
        self.category_section.pack(fill="x", padx=10, pady=5)

        self.amount_section = tk.Frame(self.root, bg="#f4a261")
        self.amount_section.pack(fill="x", padx=10, pady=5)

        self.custom_category_section = tk.Frame(self.root, bg="#f4a261")
        self.custom_category_section.pack(fill="x", padx=10, pady=5)

        # Create entry fields
        self.date_label = tk.Label(self.date_section, text="Date (DD):", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.date_label.pack(side="left")
        self.date_entry = tk.Entry(self.date_section, width=5, bg=entry_bg)
        self.date_entry.pack(side="left", padx=5)
        self.date_entry.bind("<Return>", lambda event: self.category_menu.focus())

        self.month_label = tk.Label(self.date_section, text="Month:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.month_label.pack(side="left")
        self.month_var = tk.StringVar()
        self.month_var.set(self.last_month)  # Set to last changed month
        self.month_menu = ttk.Combobox(self.date_section, textvariable=self.month_var, values=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
        self.month_menu.pack(side="left", padx=5)
        self.month_menu.bind("<<ComboboxSelected>>", self.save_last_month)

        self.year_label = tk.Label(self.date_section, text="Year:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.year_label.pack(side="left")
        self.year_var = tk.StringVar()
        self.year_var.set("2024")  # default year
        self.year_menu = ttk.Combobox(self.date_section, textvariable=self.year_var, values=["2023", "2024", "2025"])
        self.year_menu.pack(side="left", padx=5)

        self.category_label = tk.Label(self.category_section, text="Category:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.category_label.pack(side="left")
        self.category_var = tk.StringVar()
        self.category_var.set("Select Category")  # default category
        self.category_menu = ttk.Combobox(self.category_section, textvariable=self.category_var, values=["Select Category"])
        self.category_menu.pack(side="left", padx=5)
        self.category_menu.bind("<FocusIn>", self.clear_category_placeholder)
        self.category_menu.bind("<FocusOut>", self.revert_category_placeholder)
        self.category_menu.bind("<Return>", lambda event: self.amount_entry.focus())
        self.category_menu.bind("<<ComboboxSelected>>", self.focus_on_amount)

        self.add_category_button = tk.Button(self.category_section, text="Add Category", command=self.show_custom_category_entry, bg="#264653", fg="white", font=button_font)
        self.add_category_button.pack(side="left", padx=5)

        self.custom_category_label = tk.Label(self.custom_category_section, text="New Category:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.custom_category_label.pack(side="left")
        self.custom_category_entry = tk.Entry(self.custom_category_section, bg=entry_bg)
        self.custom_category_entry.pack(side="left", padx=5)
        self.custom_category_entry.bind("<Return>", self.add_category)
        self.custom_category_section.pack_forget()

        self.amount_label = tk.Label(self.amount_section, text="Amount:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.amount_label.pack(side="left")
        self.amount_entry = tk.Entry(self.amount_section, bg=entry_bg)
        self.amount_entry.pack(side="left", padx=5)
        self.amount_entry.bind("<Return>", self.submit_form)

        # Initialize workbook and filename
        self.wb = None
        self.file_name = "expenses.xlsx"
        self.load_or_create_file(self.file_name)

        # Create a dictionary to store the data month-wise and category-wise
        self.monthly_expenses = {}
        self.category_expenses = {}

        # Create a button to submit the form (after status bar)
        self.submit_button = tk.Button(self.root, text="Submit", command=self.submit_form, bg="#264653", fg="white", font=button_font)
        self.submit_button.pack(fill="x", pady=10)

        # Add Tooltips
        self.create_tooltip(self.submit_button, "Click to submit your expense entry")
        self.create_tooltip(self.add_category_button, "Add a new category")

        # Add Hover Effects
        self.submit_button.bind("<Enter>", self.on_hover)
        self.submit_button.bind("<Leave>", self.on_leave)
        self.add_category_button.bind("<Enter>", self.on_hover)
        self.add_category_button.bind("<Leave>", self.on_leave)

    def load_last_month(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as file:
                return file.read().strip()
        return "January"

    def save_last_month(self, event=None):
        with open(self.config_file, "w") as file:
            file.write(self.month_var.get())

    def get_month_number(self, month_name):
        month_numbers = {
            "January": 1,
            "February": 2,
            "March": 3,
            "April": 4,
            "May": 5,
            "June": 6,
            "July": 7,
            "August": 8,
            "September": 9,
            "October": 10,
            "November": 11,
            "December": 12
        }
        return month_numbers[month_name]

    def show_custom_category_entry(self):
        self.custom_category_section.pack(fill="x")
        self.custom_category_entry.focus()

    def clear_category_placeholder(self, event):
        if self.category_var.get() == "Select Category":
            self.category_var.set("")

    def revert_category_placeholder(self, event):
        if not self.category_var.get():
            self.category_var.set("Select Category")

    def add_category(self, event=None):
        category = self.custom_category_entry.get()
        if category and category not in self.category_menu['values']:
            self.category_menu['values'] = (*self.category_menu['values'], category)
            self.category_var.set(category)
            self.custom_category_entry.delete(0, tk.END)
            self.custom_category_section.pack_forget()
            self.amount_entry.focus()

    def focus_on_amount(self, event=None):
        self.amount_entry.focus()

    def create_new_file(self):
        file_to_open = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_to_open:
            self.file_name = file_to_open
            self.wb = Workbook()
            self.wb.save(self.file_name)
            self.update_status_bar(f"New file created: {os.path.basename(file_to_open)}")

    def load_file(self):
        file_to_open = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_to_open:
            self.file_name = file_to_open
            self.load_or_create_file(self.file_name)
            self.update_status_bar(f"File loaded: {os.path.basename(file_to_open)}")

    def load_or_create_file(self, file_name):
        if os.path.exists(file_name):
            self.wb = load_workbook(file_name)
            self.update_status_bar(f"File loaded: {os.path.basename(file_name)}")
        else:
            self.wb = Workbook()
            self.wb.save(file_name)
            self.update_status_bar(f"New file created: {os.path.basename(file_name)}")

    def save_file(self):
        self.wb.save(self.file_name)
        self.update_status_bar(f"File saved: {os.path.basename(self.file_name)}")

    def submit_form(self, event=None):
        # Get the values from the entry fields
        day = self.date_entry.get()
        month = self.month_var.get()
        year = self.year_var.get()
        date_str = f"{year}-{self.get_month_number(month)}-{day}"
        category = self.category_var.get()
        amount = float(self.amount_entry.get())

        # Parse the date string to get the month
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_name = date_obj.strftime("%B")

        # Add the data to the Excel sheet
        if category not in self.wb.sheetnames:
            ws = self.wb.create_sheet(title=category)
            ws.append(["Date", "Month", "Amount"])
        else:
            ws = self.wb[category]

        ws.append([date_str, month_name, amount])
        self.wb.save(self.file_name)

        # Clear entry fields and refocus on date entry
        self.amount_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.date_entry.focus()

        # Update the status bar
        self.update_status_bar(f"Entry added: {category} - {amount} on {date_str}")

    def update_status_bar(self, message):
        self.status_bar.config(text=message)

    def create_tooltip(self, widget, text):
        tooltip = tk.Toplevel(widget)
        tooltip.wm_overrideredirect(True)
        tooltip.wm_geometry("+%d+%d" % (widget.winfo_rootx() + 20, widget.winfo_rooty() + 20))
        label = tk.Label(tooltip, text=text, bg="yellow", relief="solid", borderwidth=1)
        label.pack()
        widget.bind("<Enter>", lambda event: tooltip.deiconify())
        widget.bind("<Leave>", lambda event: tooltip.withdraw())
        tooltip.withdraw()

    def on_hover(self, event):
        event.widget.config(bg="#e76f51")

    def on_leave(self, event):
        event.widget.config(bg="#264653")


if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
