import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")
        self.root.geometry("500x400")
        self.root.configure(bg="#f4a261")

        # Apply a theme
        style = ttk.Style()
        style.theme_use('clam')

        # Custom Fonts and Colors
        label_font = ("Helvetica", 12)
        button_font = ("Helvetica", 14, "bold")
        entry_bg = "#fff8e1"

        # Create a menu bar
        self.menu_bar = tk.Menu(self.root)
        self.root.config(menu=self.menu_bar)

        # Create the File menu
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0, bg="#e9c46a")
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)
        self.file_menu.add_command(label="New File", command=self.create_new_file)
        self.file_menu.add_command(label="Open File", command=self.load_file)
        self.file_menu.add_command(label="Save File", command=self.save_file)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Exit", command=self.root.quit)

        # Create Status Bar (early initialization)
        self.status_bar = tk.Label(self.root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W, bg="#264653", fg="white")
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Load last selected month
        self.config_file = "config.txt"
        self.last_month = self.load_last_month()

        # Create sections
        self.date_section = tk.Frame(self.root, bg="#f4a261")
        self.date_section.pack(fill="x", padx=10, pady=5)

        self.category_section = tk.Frame(self.root, bg="#f4a261")
        self.category_section.pack(fill="x", padx=10, pady=5)

        self.amount_section = tk.Frame(self.root, bg="#f4a261")
        self.amount_section.pack(fill="x", padx=10, pady=5)

        # Create entry fields
        self.date_label = tk.Label(self.date_section, text="Date (DD):", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.date_label.pack(side="left")
        self.date_entry = tk.Entry(self.date_section, width=5, bg=entry_bg)
        self.date_entry.pack(side="left", padx=5)
        self.date_entry.bind("<Return>", lambda event: self.category_menu.focus())

        self.month_label = tk.Label(self.date_section, text="Month:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.month_label.pack(side="left")
        self.month_var = tk.StringVar()
        self.month_var.set(self.last_month)  # Set to last changed month
        self.month_menu = ttk.Combobox(self.date_section, textvariable=self.month_var, values=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
        self.month_menu.pack(side="left", padx=5)
        self.month_menu.bind("<<ComboboxSelected>>", self.save_last_month)

        self.year_label = tk.Label(self.date_section, text="Year:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.year_label.pack(side="left")
        self.year_var = tk.StringVar()
        self.year_var.set("2024")  # default year
        self.year_menu = ttk.Combobox(self.date_section, textvariable=self.year_var, values=["2023", "2024", "2025"])
        self.year_menu.pack(side="left", padx=5)

        self.category_label = tk.Label(self.category_section, text="Category:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.category_label.pack(side="left")
        self.category_var = tk.StringVar()
        self.category_var.set("Select Category")  # default category
        self.category_menu = ttk.Combobox(self.category_section, textvariable=self.category_var, values=["Select Category", "Food", "Transport", "Rent", "Entertainment", "Utilities", "Miscellaneous"])
        self.category_menu.pack(side="left", padx=5)
        self.category_menu.bind("<FocusIn>", self.clear_category_placeholder)
        self.category_menu.bind("<Return>", lambda event: self.amount_entry.focus())

        self.amount_label = tk.Label(self.amount_section, text="Amount:", font=label_font, fg="#2a9d8f", bg="#f4a261")
        self.amount_label.pack(side="left")
        self.amount_entry = tk.Entry(self.amount_section, bg=entry_bg)
        self.amount_entry.pack(side="left", padx=5)
        self.amount_entry.bind("<Return>", self.submit_form)

        # Submit button
        self.submit_button = tk.Button(self.root, text="Submit", font=button_font, bg="#e76f51", fg="white", command=self.submit_form)
        self.submit_button.pack(pady=10)

        # Load or create a file
        self.file_name = "expenses.xlsx"
        self.load_or_create_file(self.file_name)

    def create_new_file(self):
        new_file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if new_file_name:
            self.file_name = new_file_name
            self.load_or_create_file(self.file_name)
            self.update_status_bar(f"New file created: {os.path.basename(new_file_name)}")

    def load_file(self):
        file_to_open = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_to_open:
            self.file_name = file_to_open
            self.load_or_create_file(self.file_name)
            self.update_status_bar(f"File loaded: {os.path.basename(file_to_open)}")

    def load_or_create_file(self, file_name):
        if os.path.exists(file_name):
            self.wb = load_workbook(file_name)
            self.update_status_bar(f"File loaded: {os.path.basename(file_name)}")
        else:
            self.wb = Workbook()
            self.wb.save(file_name)
            self.update_status_bar(f"New file created: {os.path.basename(file_name)}")

    def save_file(self):
        self.wb.save(self.file_name)
        self.update_status_bar(f"File saved: {os.path.basename(self.file_name)}")

    def update_status_bar(self, message):
        self.status_bar.config(text=message)

    def clear_category_placeholder(self, event):
        if self.category_var.get() == "Select Category":
            self.category_var.set("")

    def load_last_month(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as file:
                return file.read().strip()
        return "January"

    def save_last_month(self, event=None):
        with open(self.config_file, "w") as file:
            file.write(self.month_var.get())

    def submit_form(self, event=None):
        day = self.date_entry.get()
        month = self.month_var.get()
        year = self.year_var.get()
        date_str = f"{year}-{self.get_month_number(month)}-{day}"
        category = self.category_var.get()
        amount = float(self.amount_entry.get())

        # Check for remarks in category
        if "(" in category and ")" in category:
            remark = category[category.index("(") + 1:category.index(")")]
            category = category.split("(")[0].strip()
        else:
            remark = ""

        # Parse the date string to get the month
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        month_name = date_obj.strftime("%B")

        # Add the data to the Excel sheet
        if category not in self.wb.sheetnames:
            ws = self.wb.create_sheet(title=category)
            ws.append(["Date", "Month", "Amount", "Remark"])
        else:
            ws = self.wb[category]

        ws.append([date_str, month_name, amount, remark])
        self.wb.save(self.file_name)

        # Clear entry fields and refocus on date entry
        self.amount_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.date_entry.focus()

        # Update the status bar
        self.update_status_bar(f"Entry added: {category} - {amount} on {date_str}")

    def get_month_number(self, month_name):
        month_numbers = {
            "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, 
            "June": 6, "July": 7, "August": 8, "September": 9, "October": 10, 
            "November": 11, "December": 12
        }
        return month_numbers[month_name]

if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
