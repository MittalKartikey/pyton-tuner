import tkinter as tk  # Import tkinter as tk
from tkinter import ttk, messagebox, filedialog  # Import ttk and other required modules from tkinter
from openpyxl import Workbook, load_workbook
import os

class ExpenseTracker:
    def __init__(self, root):
        self.root = root
        self.root.title("Expense Tracker")
        self.root.geometry("400x250")
        self.root.configure(bg="black")

        # Apply a theme
        style = ttk.Style()
        style.theme_use('clam')

        # Custom Fonts and Colors
        label_font = ("Helvetica", 12, "bold")
        button_font = ("Helvetica", 14, "bold")
        entry_bg = "white"
        entry_fg = "black"
        entry_width = 20  # Consistent width for all entries

        # Create a menu bar
        self.menu_bar = tk.Menu(self.root, bg="#264653", fg="white")
        self.root.config(menu=self.menu_bar)

        # Create the File menu
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0, bg="#e9c46a", fg="black")
        self.menu_bar.add_cascade(label="File", menu=self.file_menu)
        self.file_menu.add_command(label="New File", command=self.create_new_file)
        self.file_menu.add_command(label="Open File", command=self.load_file)
        self.file_menu.add_command(label="Save File", command=self.save_file)
        self.file_menu.add_separator()
        self.file_menu.add_command(label="Exit", command=self.root.quit)

        # Create Status Bar
        self.status_bar = tk.Label(self.root, text="Ready", bd=1, relief=tk.SUNKEN, anchor=tk.W, bg="#264653", fg="white")
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Load last selected month
        self.config_file = "config.txt"
        self.last_month = self.load_last_month()

        # Create form grid layout
        form_frame = tk.Frame(self.root, bg="black")
        form_frame.pack(padx=10, pady=10, anchor="w")

        # Create entry fields with consistent width
        self.date_label = tk.Label(form_frame, text="Date (DD):", font=label_font, fg="white", bg="black")
        self.date_label.grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.date_entry = ttk.Entry(form_frame, width=entry_width, style='RoundedEntry.TEntry')
        self.date_entry.grid(row=0, column=1, padx=5, pady=5)
        self.date_entry.bind("<Return>", lambda event: self.category_menu.focus())

        self.month_label = tk.Label(form_frame, text="Month:", font=label_font, fg="white", bg="black")
        self.month_label.grid(row=1, column=0, sticky="e", padx=5, pady=5)
        self.month_var = tk.StringVar()
        self.month_var.set(self.last_month)
        self.month_menu = ttk.Combobox(form_frame, textvariable=self.month_var, values=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"], width=entry_width-1)
        self.month_menu.grid(row=1, column=1, padx=5, pady=5)
        self.month_menu.bind("<<ComboboxSelected>>", self.save_last_month)

        self.year_label = tk.Label(form_frame, text="Year:", font=label_font, fg="white", bg="black")
        self.year_label.grid(row=2, column=0, sticky="e", padx=5, pady=5)
        self.year_var = tk.StringVar()
        self.year_var.set("2024")
        self.year_menu = ttk.Combobox(form_frame, textvariable=self.year_var, values=["2023", "2024", "2025"], width=entry_width-1)
        self.year_menu.grid(row=2, column=1, padx=5, pady=5)

        self.category_label = tk.Label(form_frame, text="Category:", font=label_font, fg="white", bg="black")
        self.category_label.grid(row=3, column=0, sticky="e", padx=5, pady=5)
        self.category_var = tk.StringVar()
        self.category_var.set("Select Category")
        self.category_menu = ttk.Combobox(form_frame, textvariable=self.category_var, values=["Select Category", "Food", "Transport", "Rent", "Entertainment", "Utilities", "Miscellaneous"], width=entry_width-1)
        self.category_menu.grid(row=3, column=1, padx=5, pady=5)
        self.category_menu.bind("<FocusIn>", self.clear_category_placeholder)
        self.category_menu.bind("<Return>", lambda event: self.amount_entry.focus())

        self.amount_label = tk.Label(form_frame, text="Amount:", font=label_font, fg="white", bg="black")
        self.amount_label.grid(row=4, column=0, sticky="e", padx=5, pady=5)
        self.amount_entry = ttk.Entry(form_frame, width=entry_width, style='RoundedEntry.TEntry')
        self.amount_entry.grid(row=4, column=1, padx=5, pady=5)
        self.amount_entry.bind("<Return>", self.submit_form)

        # Submit button with reduced padding
        self.submit_button = tk.Button(self.root, text="Submit", font=button_font, bg="blue", fg="white", command=self.submit_form)
        self.submit_button.pack(pady=5)  # Reduced padding

        # Load or create a file
        self.file_name = "expenses.xlsx"
        self.load_or_create_file(self.file_name)

        # Style configuration for rounded entries
        style.configure('RoundedEntry.TEntry', fieldbackground=entry_bg, foreground=entry_fg, borderwidth=1)
        style.map('RoundedEntry.TEntry', background=[('focus', '#264653')])

    def create_new_file(self):
        new_file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if new_file_name:
            self.file_name = new_file_name
            self.load_or_create_file(self.file_name)
            self.update_status_bar(f"New file created: {os.path.basename(new_file_name)}")

    def load_file(self):
        file_to_open = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_to_open:
            self.file_name = file_to_open
            self.load_or_create_file(self.file_name)
            self.update_status_bar(f"File loaded: {os.path.basename(file_to_open)}")

    def load_or_create_file(self, file_name):
        if os.path.exists(file_name):
            self.wb = load_workbook(file_name)
            self.update_status_bar(f"File loaded: {os.path.basename(file_name)}")
        else:
            self.wb = Workbook()
            self.wb.save(file_name)
            self.update_status_bar(f"New file created: {os.path.basename(file_name)}")

    def save_file(self):
        self.wb.save(self.file_name)
        self.update_status_bar(f"File saved: {os.path.basename(self.file_name)}")

    def update_status_bar(self, message):
        self.status_bar.config(text=message)

    def clear_category_placeholder(self, event):
        if self.category_var.get() == "Select Category":
            self.category_var.set("")

    def load_last_month(self):
        if os.path.exists(self.config_file):
            with open(self.config_file, "r") as file:
                return file.read().strip()
        return "January"

    def save_last_month(self, event=None):
        with open(self.config_file, "w") as file:
            file.write(self.month_var.get())

    def submit_form(self, event=None):
        day = self.date_entry.get()
        month = self.month_var.get()
        year = self.year_var.get()
        date_str = f"{year}-{self.get_month_number(month)}-{day}"
        category = self.category_var.get()
        amount = float(self.amount_entry.get())

        if category not in self.wb.sheetnames:
            ws = self.wb.create_sheet(title=category)
            ws.append(["Date", "Month", "Amount", "Remark"])
        else:
            ws = self.wb[category]

        # Add an entry to the sheet
        ws.append([date_str, month, amount, ""])
        self.wb.save(self.file_name)

        # Clear entry fields and refocus on date entry
        self.amount_entry.delete(0, tk.END)
        self.date_entry.delete(0, tk.END)
        self.date_entry.focus()

        self.update_status_bar(f"Entry added: {category} - {amount} on {date_str}")

    def get_month_number(self, month_name):
        month_numbers = {
            "January": 1, "February": 2, "March": 3, "April": 4, "May": 5, 
            "June": 6, "July": 7, "August": 8, "September": 9, "October": 10,
            "November": 11, "December": 12
        }
        return month_numbers[month_name]

if __name__ == "__main__":
    root = tk.Tk()
    app = ExpenseTracker(root)
    root.mainloop()
